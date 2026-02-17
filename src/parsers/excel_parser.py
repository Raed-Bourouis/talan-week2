"""Excel document parser."""
import logging
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel documents."""
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse an Excel document and extract structured data."""
        try:
            text = self.extract_text(file_path)
            sheets_data = self._extract_sheets_data(file_path)
            
            return {
                "text": text,
                "sheets": sheets_data,
                "file_type": "excel"
            }
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            raise
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from an Excel document."""
        try:
            workbook = load_workbook(file_path, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
                
                text_parts.append("")  # Empty line between sheets
            
            return "\n".join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from Excel {file_path}: {e}")
            raise
    
    def _extract_sheets_data(self, file_path: str) -> List[Dict[str, Any]]:
        """Extract structured data from all sheets."""
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheets_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get all rows as list of lists
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append([cell if cell is not None else "" for cell in row])
                
                sheets_data.append({
                    "name": sheet_name,
                    "data": data,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column
                })
            
            return sheets_data
        except Exception as e:
            logger.warning(f"Error extracting sheets data from Excel {file_path}: {e}")
            return []
